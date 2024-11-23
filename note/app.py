from flask import Flask, render_template, request, redirect, url_for, session, flash
import openpyxl
import os

app = Flask(__name__)
app.secret_key = 'secret_key'

# Classes avec prénoms triés par ordre alphabétique
classes = {
    '3°1': sorted(['Alice', 'Bob', 'Cédric']),
    '3°2': sorted(['David', 'Eve', 'François']),
    '3°3': sorted(['Julia 3°3', 'Agathe 3°3', 'Gaïa 3°3', 'Chloé 3°3', 'Coleen 3°3', 'Klervi 3°3', 'Moyra 3°3', 'Anaïs P 3°3', 'Anaïs U 3°3', 'Ines 3°3']),
    '3°4': sorted(['Charlotte 3°4', 'Ninon 3°4', 'Emilie A 3°4', 'Emilie C 3°4', 'Victoria 3°4', 'Clemence 3°4', 'Roman 3°4', 'Mathilde 3°4', 'Manon C 3°4', 'Lena 3°4', 'Zoe 3°4']),
    'Special': ['Special']  # Classe spéciale pour sélection
}

# Dictionnaire des utilisateurs avec leurs mots de passe
users_passwords = {
    'Benjamin Hickey': '30MD',
    'Paul Dubesset': '29JP',
    'Gabriel Cornet': '28EA',
    'Gicquel Amaury': '30GT',
    'Jules Parienté': '13MDV'
}

# Fichier Excel pour les résultats
file_path = 'special_results.xlsx'

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username in users_passwords and users_passwords[username] == password:
            session['username'] = username
            return redirect(url_for('class_selection'))
        else:
            flash('Nom d\'utilisateur ou mot de passe incorrect', 'error')

    return render_template('login.html')

@app.route('/class_selection', methods=['GET', 'POST'])
def class_selection():
    if request.method == 'POST':
        selected_class = request.form['class']
        session['selected_class'] = selected_class
        
        if selected_class == 'Special':
            return redirect(url_for('special_class_selection'))
        else:
            session['students'] = classes[selected_class]
            session['current_student'] = 0
            session['grades'] = []
            return redirect(url_for('rate_student'))

    return render_template('class_selection.html', classes=classes.keys())

@app.route('/special_class_selection', methods=['GET', 'POST'])
def special_class_selection():
    if request.method == 'POST':
        selected_class = request.form['class']
        session['selected_special_class'] = selected_class
        session['students'] = classes[selected_class]
        session['current_student'] = 0
        session['grades'] = []
        return redirect(url_for('rate_special_student'))

    return render_template('special_class_selection.html', classes=list(classes.keys())[0:-1])  # Exclure la classe spéciale de la sélection

@app.route('/rate_student', methods=['GET', 'POST'])
def rate_student():
    current_student = session['current_student']
    if request.method == 'POST':
        grade = request.form['grade']
        session['grades'].append((session['students'][current_student], grade))
        session['current_student'] += 1
        if session['current_student'] >= len(session['students']):
            save_results(session['username'], session['selected_class'], session['grades'], file_path)
            return redirect(url_for('result'))
    return render_template('rating.html', student=session['students'][current_student])

@app.route('/rate_special_student', methods=['GET', 'POST'])
def rate_special_student():
    current_student = session['current_student']
    if request.method == 'POST':
        grade_tete = request.form['grade_tete']
        grade_sein = request.form['grade_sein']
        grade_fesse = request.form['grade_fesse']
        grade_personnalite = request.form['grade_personnalite']
        
        total_grade = (float(grade_tete) + float(grade_sein) + float(grade_fesse) + float(grade_personnalite)) * 4
        
        session['grades'].append((session['students'][current_student], grade_tete, grade_sein, grade_fesse, grade_personnalite, total_grade))
        session['current_student'] += 1
        if session['current_student'] >= len(session['students']):
            save_special_results(session['username'], session['selected_special_class'], session['grades'])
            return redirect(url_for('result'))
    return render_template('rating_special.html', student=session['students'][current_student])

@app.route('/result', methods=['GET', 'POST'])
def result():
    current_user = session.get('username')

    if request.method == 'POST':
        return redirect(url_for('index'))

    special_grades = load_all_results()

    # Filtrer les résultats en fonction de l'utilisateur
    if current_user == "Jules Parienté":
        filtered_special_grades = special_grades
    else:
        filtered_special_grades = [grade for grade in special_grades if grade[0] == current_user]

    return render_template('result.html', special_grades=filtered_special_grades)

def save_special_results(username, selected_class, grades):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Special Results'
        ws.append(['Utilisateur', 'Classe', 'Fille', 'Tête', 'Seins', 'Fesses', 'Personnalité', 'Total'])
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

    for student, grade_tete, grade_sein, grade_fesse, grade_personnalite, total_grade in grades:
        ws.append([username, selected_class, student, grade_tete, grade_sein, grade_fesse, grade_personnalite, total_grade])

    wb.save(file_path)

def load_all_results():
    if not os.path.exists(file_path):
        return []
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    all_grades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        all_grades.append(row)

    return all_grades

if __name__ == '__main__':
    app.run(debug=True)
